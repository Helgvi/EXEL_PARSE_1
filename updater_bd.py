import openpyxl
import sqlite3


con = sqlite3.connect('db.sqlite')
cur = con.cursor()


def create_bd():
    cur.execute('''
      CREATE TABLE IF NOT EXISTS goods(
        id INTEGER PRIMARY KEY,
        Ccode INTEGER,
        Vcode TEXT,
        art TEXT,
        part TEXT,
        price INTEGER,
        inv TEXT,
        quel INTEGER
        );
    ''')


def import_table_to_bd(index):
    cur = con.cursor()
    wb = openpyxl.load_workbook(filename='c:\Price2.xlsx', read_only=True)
    list_name = wb.sheetnames
    sheet = wb[list_name[0]]
    i = index
    for row in sheet.iter_rows(min_row=1,
                               max_row=22519,
                               min_col=1,
                               max_col=7,
                               values_only=True):
        i = i + 1
        line = (i, row[0], row[1], row[2], row[3], row[4], row[5], row[6])
        print(line)
        cur.execute(
            'INSERT INTO goods VALUES(?, ?, ?, ?, ?, ?, ?, ?);',
            line
        )


create_bd()
import_table_to_bd(0)

con.commit()
con.close()
