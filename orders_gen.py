import sqlite3
import requests
import openpyxl
import xlwt


VOSHOD_TOKEN = '5025-AjsXKgYEY9QCXkEatwCWFBrdjjEwheC8jUDZUEfxN3hFx83GmA3nbckjKHcKrqWjwY9ggwJdAcCbzt5P'
PATH_EXTAKE = 'c:/Intake/orders.xls'
intake_list = {}
ENDPOINT = 'https://api.v-avto.ru/v1/search/name.json?'
HEADERS = {'X-Voshod-API-KEY': VOSHOD_TOKEN}
CHECK_DICT = 'Запрос вернул не словарь! {}'
FAIL_CONNECT = 'Запрос отклонен {}, Запрос к "{url}", {headers}, {params}.'
NOT_CODE = '-'
Ccode = list()


def response(context):
    """API запрос к базе Восход"""
    params_for_response = dict(
        url=ENDPOINT,
        headers=HEADERS,
        params={'q': context}
    )
    try:
        response = requests.get(**params_for_response)
    except requests.exceptions.RequestException as error:
        raise ConnectionError(
            FAIL_CONNECT.format(error, **params_for_response)
        )
    if response.status_code == 200:
        answer_api = response.json()
    if not isinstance(response.json(), dict):
        raise TypeError(CHECK_DICT.format(type(response)))
    answer = answer_api.get('response').get('page').get('items') 
    if answer == 0:
        counts = NOT_CODE
    elif answer == 1:
        counts = answer_api.get('response').get('items')[0].get('count')
    return counts


def need_count(count_post, your_count):
    if your_count == ' ':
        your_count = 0
    if count_post == 0 and float(your_count) >= 0:
        return 'НЕТ У поставщика'
    elif count_post == NOT_CODE:
        return NOT_CODE
    elif int(count_post) > 0 and float(your_count) >= 0:
        return 3


def bild_list():
    "Список всех кодов из базы данных"
    con = sqlite3.connect('db.sqlite')
    cur = con.cursor()
    sql = 'SELECT Vcode, Ccode FROM goods;'
    cur.execute(sql)
    for result in cur:
        intake_list[result[1]] = result[0]
    con.commit()
    con.close()
    for key in intake_list.keys():
        Ccode.append(key)
    return intake_list


def answer_from_exel_file(path_intake):
    "Обработка входящих данных исходной таблицы"
    wb = openpyxl.load_workbook(filename=path_intake, read_only=True)
    list_name = wb.sheetnames
    sheet = wb[list_name[0]]
    index = 0
    book = xlwt.Workbook(encoding="utf-8")
    sheet1 = book.add_sheet("Накладная")
    for row in sheet.iter_rows(min_row=1,
                               max_row=100,
                               min_col=1,
                               max_col=10,
                               values_only=True):
        index = index + 1
        if row[0] not in Ccode:
            res = 'НЕТ Кода'
        else:
            if intake_list[row[0]] == ' ':
                res = 'НЕТ Кода'
            else:
                res = intake_list[row[0]]
                count = response(res)
        sheet1.write(index, 0, row[0])
        sheet1.write(index, 1, res)
        sheet1.write(index, 2, row[1])
        sheet1.write(index, 3, row[2])
        sheet1.write(index, 4, row[3])
        sheet1.write(index, 5, row[4])
        sheet1.write(index, 6, row[5])
        sheet1.write(index, 7, row[6])
        sheet1.write(index, 8, count)
        sheet1.write(index, 9, need_count(count, row[5]))
        book.save(PATH_EXTAKE)


def main(path_intake):
    bild_list()
    answer_from_exel_file(path_intake)
    return 'Выполнено!'


if __name__ == '__main__':
    main()
